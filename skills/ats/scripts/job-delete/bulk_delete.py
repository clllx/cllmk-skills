"""Bulk hard-delete Moka ATS jobs via cllmk curl.

Reads jobIds from xlsx/csv or --job-id flags, calls
POST /api/outer/ats-jc/job/jobs/deleteJob once per ID (backend is single-ID),
classifies failures without retrying result-unknown network errors, writes report.xlsx.

Default is preview-only; pass --confirm to actually delete (hard delete, irreversible).

Usage:
  python3 bulk_delete.py --input <xlsx|csv> [--id-column <name|index>]
                        [--workdir <dir>] [--interval 1.5]
                        [--confirm]
  python3 bulk_delete.py --job-id <uuid> [--job-id <uuid> ...] --workdir <dir>
                        [--confirm]

Resume: re-running with the same --workdir picks up from state.json.
"""
from __future__ import annotations

import argparse
import csv
import json
import os
import re
import subprocess
import sys
import time
import uuid
from datetime import datetime
from pathlib import Path

URL = "/api/outer/ats-jc/job/jobs/deleteJob"
ORG_INFO_URL = "/api/v2/org/info"
NETWORK_FAIL_PATTERN = re.compile(
    r"Client network socket disconnected|ECONNRESET|ETIMEDOUT|socket hang up|EAI_AGAIN",
    re.I,
)
KNOWN_BIZ_CODES = {
    704023: "职位下有招聘中候选人",
    704024: "有关联职位（需先解除关联）",
    705004: "职位不存在（可视为已完成）",
}


def require_current_tenant() -> None:
    if os.environ.get("CLLMK_PROFILE", "").strip():
        sys.exit("error: CLLMK_PROFILE is set; unset it, run 'cllmk auth switch', "
                 "verify with 'cllmk auth status', then retry")


def is_uuid(s: str) -> bool:
    try:
        uuid.UUID(str(s).strip())
        return True
    except (ValueError, AttributeError):
        return False


def load_ids_from_file(path: Path, id_column) -> list[str]:
    if path.suffix.lower() in {".xlsx", ".xlsm"}:
        try:
            import openpyxl
        except ImportError:
            sys.exit("error: openpyxl required to read xlsx. pip install openpyxl")
        wb = openpyxl.load_workbook(path, read_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
    elif path.suffix.lower() == ".csv":
        with path.open(newline="") as f:
            rows = list(csv.reader(f))
    else:
        sys.exit(f"error: unsupported input format {path.suffix}")

    if not rows:
        sys.exit("error: input file is empty")

    header = rows[0]
    if isinstance(id_column, str) and not id_column.isdigit():
        try:
            col_idx = list(header).index(id_column)
        except ValueError:
            sys.exit(f"error: column '{id_column}' not found in header {header}")
        data_rows = rows[1:]
    else:
        col_idx = int(id_column) if id_column is not None else 0
        first_val = rows[0][col_idx] if col_idx < len(rows[0]) else None
        looks_like_header = first_val is not None and not is_uuid(str(first_val))
        data_rows = rows[1:] if looks_like_header else rows

    ids: list[str] = []
    for row in data_rows:
        if col_idx >= len(row):
            continue
        v = row[col_idx]
        if v is None or str(v).strip() == "":
            continue
        ids.append(str(v).strip())
    return ids


def fetch_org_id() -> str:
    """Return current cllmk session's orgId. Deletion is org-scoped and
    irreversible, so an undeterminable org aborts instead of proceeding."""
    require_current_tenant()
    proc = subprocess.run(
        ["cllmk", "curl", "--url", ORG_INFO_URL, "--method", "GET"],
        capture_output=True, text=True, timeout=60,
    )
    if proc.returncode != 0:
        sys.exit(f"error: cllmk curl org/info failed rc={proc.returncode} "
                 f"err={proc.stderr.strip()[:200]} out={proc.stdout.strip()[:200]}")
    try:
        resp = json.loads(proc.stdout)
    except json.JSONDecodeError:
        sys.exit(f"error: org/info returned non-JSON: {proc.stdout.strip()[:200]}")
    if resp.get("code") != 0:
        sys.exit(f"error: org/info code={resp.get('code')} msg={resp.get('msg')}")
    data = resp.get("data") or {}
    org_id = (data.get("orgInfo") or {}).get("orgId")
    if not org_id:
        sys.exit("error: org/info response has no orgInfo.orgId; refuse to delete "
                 "without knowing the target org")
    return str(org_id)


def require_expected_org(org_id: str, expected_org_id: str | None) -> None:
    if not expected_org_id:
        sys.exit("error: --expected-org-id is required with --confirm")
    if org_id != expected_org_id:
        sys.exit(f"error: current cllmk orgId={org_id} does not match "
                 f"--expected-org-id={expected_org_id}; refuse to delete jobs")


def call_delete(job_id: str) -> tuple[int, str, str]:
    require_current_tenant()
    payload = json.dumps({"jobId": job_id, "clientType": "main"})
    proc = subprocess.run(
        ["cllmk", "curl", "--method", "POST", "--url", URL, "--payload", payload],
        capture_output=True,
        text=True,
        timeout=60,
    )
    return proc.returncode, proc.stdout.strip(), proc.stderr.strip()


def classify(rc: int, out: str, err: str) -> tuple[str, dict]:
    """Return (status, meta) where meta may contain errorCode/msg/opNo.

    status in {OK, NETWORK_FAIL, BUSINESS_FAIL, OTHER_FAIL}.
    """
    if rc != 0:
        combined = (err or "") + "\n" + (out or "")
        if NETWORK_FAIL_PATTERN.search(combined):
            return "NETWORK_FAIL", {"detail": combined[:300]}
        return "OTHER_FAIL", {"detail": (err or out)[:300]}

    if not out:
        return "OTHER_FAIL", {"detail": "empty response"}

    try:
        resp = json.loads(out)
    except json.JSONDecodeError:
        return "OTHER_FAIL", {"detail": out[:300]}

    inner = resp.get("data") if isinstance(resp.get("data"), dict) else {}
    inner_msg = (inner.get("msg") if inner else None) or resp.get("msg") or ""

    if resp.get("code") == 0 and inner and inner.get("success") is True:
        return "OK", {"detail": inner_msg}

    if NETWORK_FAIL_PATTERN.search(json.dumps(resp, ensure_ascii=False)):
        return "NETWORK_FAIL", {"detail": inner_msg[:300]}

    if inner and inner.get("success") is False:
        return "BUSINESS_FAIL", {
            "errorCode": inner.get("code"),
            "detail": inner_msg,
            "opNo": inner.get("opNo", ""),
        }

    return "OTHER_FAIL", {"detail": inner_msg or resp.get("msg", "")[:300]}


def make_logger(log_path: Path):
    def log(msg: str):
        line = f"[{datetime.now().isoformat(timespec='seconds')}] {msg}"
        print(line, flush=True)
        with log_path.open("a") as f:
            f.write(line + "\n")

    return log


def write_report(workdir: Path, records: list[dict], stem: str = "report"):
    """Write report.xlsx if openpyxl available; else report.csv."""
    columns = ["jobId", "status", "errorCode", "detail", "opNo", "attempts"]
    try:
        import openpyxl

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "report"
        ws.append(columns)
        for r in records:
            ws.append([r.get(c, "") for c in columns])
        wb.save(workdir / f"{stem}.xlsx")
        return workdir / f"{stem}.xlsx"
    except ImportError:
        report_path = workdir / f"{stem}.csv"
        with report_path.open("w", newline="") as f:
            w = csv.writer(f)
            w.writerow(columns)
            for r in records:
                w.writerow([r.get(c, "") for c in columns])
        return report_path


def process(
    job_ids: list[str],
    workdir: Path,
    interval: float,
    dry_run: bool,
    org_id: str = "",
) -> list[dict]:
    log_name = "preview.log" if dry_run else "delete.log"
    log = make_logger(workdir / log_name)
    state_path = workdir / "state.json"
    records: list[dict] = []

    start_idx = 0
    # Preview runs must not consume or advance live resume state. Otherwise a
    # later --confirm run in the same workdir can incorrectly skip every ID.
    if not dry_run and state_path.exists():
        try:
            st = json.loads(state_path.read_text())
        except Exception:
            st = None
        if st:
            if org_id and st.get("orgId") and st["orgId"] != org_id:
                sys.exit(f"error: state.json orgId={st['orgId']} but current cllmk org={org_id}; "
                         f"delete {state_path} or switch cllmk profile/env before rerunning")
            start_idx = st.get("next_index", 0)
            records = st.get("records", [])
            if start_idx > 0:
                log(f"RESUME from index {start_idx} (records so far: {len(records)})")

    mode = "DRY-RUN" if dry_run else "LIVE"
    log(f"{mode} START total={len(job_ids)} interval={interval}s")

    counts = {"OK": 0, "BUSINESS_FAIL": 0, "NETWORK_FAIL": 0, "OTHER_FAIL": 0, "SKIPPED_INVALID": 0}
    for c in [r.get("status", "") for r in records]:
        if c in counts:
            counts[c] += 1

    for idx in range(start_idx, len(job_ids)):
        jid = job_ids[idx]

        if not is_uuid(jid):
            rec = {
                "jobId": jid,
                "status": "SKIPPED_INVALID",
                "errorCode": "",
                "detail": "not a UUID",
                "opNo": "",
                "attempts": 0,
            }
            records.append(rec)
            counts["SKIPPED_INVALID"] += 1
            log(f"{idx+1}/{len(job_ids)} SKIPPED_INVALID id={jid}")
            if not dry_run:
                state_path.write_text(json.dumps({"next_index": idx + 1, "records": records, "orgId": org_id}))
            continue

        if dry_run:
            rec = {
                "jobId": jid,
                "status": "DRY_RUN",
                "errorCode": "",
                "detail": "would delete",
                "opNo": "",
                "attempts": 0,
            }
            records.append(rec)
            log(f"{idx+1}/{len(job_ids)} DRY_RUN id={jid}")
            continue

        # A network failure may occur after the server committed the delete.
        # Never retry automatically; record the result as unknown for verification.
        attempts = 0
        status = "OTHER_FAIL"
        meta: dict = {}
        attempts += 1
        try:
            rc, out, err = call_delete(jid)
            status, meta = classify(rc, out, err)
        except Exception as e:
            status, meta = "NETWORK_FAIL", {"detail": f"EXC {e}"}
        if status == "NETWORK_FAIL":
            meta["detail"] = (meta.get("detail", "") +
                              " | result unknown; verify job existence before manual retry")

        rec = {
            "jobId": jid,
            "status": status,
            "errorCode": meta.get("errorCode", ""),
            "detail": meta.get("detail", "")[:300],
            "opNo": meta.get("opNo", ""),
            "attempts": attempts,
        }
        records.append(rec)
        counts[status] = counts.get(status, 0) + 1

        if status == "OK":
            log(f"{idx+1}/{len(job_ids)} OK id={jid} attempts={attempts}")
        elif status == "BUSINESS_FAIL":
            hint = KNOWN_BIZ_CODES.get(meta.get("errorCode"), "unknown")
            log(
                f"{idx+1}/{len(job_ids)} BUSINESS_FAIL id={jid} code={meta.get('errorCode')} "
                f"hint={hint} msg={meta.get('detail','')[:120]}"
            )
        else:
            log(f"{idx+1}/{len(job_ids)} {status} id={jid} attempts={attempts} msg={meta.get('detail','')[:120]}")

        state_path.write_text(json.dumps({"next_index": idx + 1, "records": records, "orgId": org_id}))

        if idx < len(job_ids) - 1:
            time.sleep(interval)

    log(f"{mode} DONE " + " ".join(f"{k}={v}" for k, v in counts.items()))
    return records


def main():
    ap = argparse.ArgumentParser(description="Bulk hard-delete Moka ATS jobs (campus + social).")
    src = ap.add_mutually_exclusive_group(required=True)
    src.add_argument("--input", type=Path, help="xlsx or csv file with jobIds")
    src.add_argument("--job-id", dest="job_ids", action="append", help="single jobId (UUID); repeatable")
    ap.add_argument("--id-column", default=None, help="column name or 0-based index (default: first column)")
    ap.add_argument("--workdir", type=Path, default=None, help="output directory (default: input file's parent dir)")
    ap.add_argument("--interval", type=float, default=1.5, help="inter-call delay in seconds (default: 1.5)")
    ap.add_argument("--confirm", action="store_true", help="disable dry-run; actually delete (irreversible)")
    ap.add_argument("--expected-org-id",
                    help="required with --confirm; must exactly match the live current orgId")
    args = ap.parse_args()

    if args.input:
        ids = load_ids_from_file(args.input, args.id_column)
        workdir = args.workdir or args.input.parent
    else:
        ids = args.job_ids
        if not args.workdir:
            sys.exit("error: --workdir required when using --job-id")
        workdir = args.workdir

    # Dedupe preserving order
    seen: set[str] = set()
    deduped: list[str] = []
    for j in ids:
        if j not in seen:
            seen.add(j)
            deduped.append(j)
    if len(deduped) != len(ids):
        print(f"note: deduped {len(ids)} → {len(deduped)} jobIds")
    ids = deduped

    workdir.mkdir(parents=True, exist_ok=True)
    if not ids:
        sys.exit("error: no jobIds found")

    dry_run = not args.confirm
    org_id = ""
    if not dry_run:
        org_id = fetch_org_id()
        require_expected_org(org_id, args.expected_org_id)
        print(f"org={org_id} (tenant_source=current)")
    est_seconds = len(ids) * (args.interval + 0.5)
    banner = "🔥 LIVE DELETE" if not dry_run else "🧪 DRY-RUN (nothing will be deleted)"
    print(
        f"{banner} | {len(ids)} jobIds | interval={args.interval}s | "
        f"est. {est_seconds:.0f}s | workdir={workdir}"
    )
    if dry_run:
        print("[preview] no API called; rerun with --confirm --expected-org-id <orgId>")

    records = process(ids, workdir, args.interval, dry_run, org_id)
    report_stem = "preview-report" if dry_run else "report"
    report_path = write_report(workdir, records, report_stem)
    print(f"report → {report_path}")


if __name__ == "__main__":
    main()
